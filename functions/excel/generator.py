"""
Excel File Generator - Create custom Excel workbooks from user requests
Supports various templates and customization options
"""

import os
import json
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, LineChart, Reference

def run(content, title="Generated Workbook", template="basic", output_path=None, **kwargs):
    """
    Generate Excel workbook based on user request
    
    Args:
        content: Main content (dict, list, or structured data)
        title: Workbook title/name
        template: Template type ('basic', 'report', 'financial', 'inventory', 'dashboard', 'budget')
        output_path: Where to save (default: reports/generated_TIMESTAMP.xlsx)
        **kwargs: Additional customization options
            - color: Primary color as hex (default: '667eea')
            - sheets: List of sheet dictionaries for multi-sheet workbooks
            - include_charts: Boolean to include charts (default: True)
            - freeze_panes: Boolean to freeze header row (default: True)
            - auto_filter: Boolean to add auto-filter (default: True)
    
    Returns:
        Path to generated Excel file
    
    Examples:
        # Simple data table
        <run:excel/generator content='{"Name": ["John", "Jane"], "Age": [30, 25]}' 
                            title="Employee List">
        
        # Financial report
        <run:excel/generator content='{"Q1": 10000, "Q2": 12000, "Q3": 15000}' 
                            title="Sales Report" 
                            template="financial">
        
        # Multi-sheet workbook
        <run:excel/generator sheets='[{"name": "Sales", "data": {...}}, 
                                      {"name": "Expenses", "data": {...}}]'
                            title="Annual Report"
                            template="report">
    """
    
    # Set default output path
    if output_path is None:
        os.makedirs('reports', exist_ok=True)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_path = f'reports/{title.lower().replace(" ", "_")}_{timestamp}.xlsx'
    else:
        os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
    
    # Get template generator
    templates = {
        'basic': generate_basic_workbook,
        'report': generate_report_workbook,
        'financial': generate_financial_workbook,
        'inventory': generate_inventory_workbook,
        'dashboard': generate_dashboard_workbook,
        'budget': generate_budget_workbook
    }
    
    generator = templates.get(template, generate_basic_workbook)
    
    # Generate workbook
    wb = generator(content, title, **kwargs)
    
    # Save file
    wb.save(output_path)
    
    return f"✅ Excel file created: {output_path}"


def generate_basic_workbook(content, title, color='667eea', freeze_panes=True, 
                            auto_filter=True, include_charts=False, sheets=None, **kwargs):
    """Generate basic Excel workbook"""
    
    wb = Workbook()
    
    if sheets:
        # Multi-sheet workbook
        if isinstance(sheets, str):
            try:
                sheets = json.loads(sheets)
            except:
                sheets = [{"name": "Sheet1", "data": content}]
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        
        for i, sheet_data in enumerate(sheets):
            ws = wb.create_sheet(title=sheet_data.get('name', f'Sheet{i+1}'))
            data = sheet_data.get('data', {})
            populate_sheet(ws, data, color, freeze_panes, auto_filter)
            
            if include_charts and isinstance(data, dict):
                add_basic_chart(ws, data)
    else:
        # Single sheet
        ws = wb.active
        ws.title = "Data"
        populate_sheet(ws, content, color, freeze_panes, auto_filter)
        
        if include_charts and isinstance(content, dict):
            add_basic_chart(ws, content)
    
    return wb


def generate_report_workbook(content, title, color='667eea', **kwargs):
    """Generate professional report workbook"""
    
    wb = Workbook()
    
    # Summary sheet
    summary_ws = wb.active
    summary_ws.title = "Summary"
    create_summary_sheet(summary_ws, title, content, color)
    
    # Data sheet
    data_ws = wb.create_sheet(title="Data")
    populate_sheet(data_ws, content, color, True, True)
    
    # Charts sheet
    if isinstance(content, dict) and kwargs.get('include_charts', True):
        charts_ws = wb.create_sheet(title="Charts")
        create_charts_sheet(charts_ws, content, color)
    
    return wb


def generate_financial_workbook(content, title, color='2E86AB', **kwargs):
    """Generate financial report workbook"""
    
    wb = Workbook()
    
    # Income Statement
    income_ws = wb.active
    income_ws.title = "Income Statement"
    create_financial_sheet(income_ws, content, "Income Statement", color)
    
    # Summary with calculations
    summary_ws = wb.create_sheet(title="Financial Summary")
    create_financial_summary(summary_ws, content, color)
    
    # Charts
    if kwargs.get('include_charts', True):
        charts_ws = wb.create_sheet(title="Analysis")
        create_financial_charts(charts_ws, content, color)
    
    return wb


def generate_inventory_workbook(content, title, color='4A5859', **kwargs):
    """Generate inventory tracking workbook"""
    
    wb = Workbook()
    
    # Main inventory sheet
    inventory_ws = wb.active
    inventory_ws.title = "Inventory"
    create_inventory_sheet(inventory_ws, content, color)
    
    # Low stock alert sheet
    alert_ws = wb.create_sheet(title="Low Stock Alert")
    create_alert_sheet(alert_ws, content, color)
    
    # Summary statistics
    stats_ws = wb.create_sheet(title="Statistics")
    create_inventory_stats(stats_ws, content, color)
    
    return wb


def generate_dashboard_workbook(content, title, color='FF6B35', **kwargs):
    """Generate dashboard-style workbook with KPIs"""
    
    wb = Workbook()
    
    # KPI Dashboard
    dashboard_ws = wb.active
    dashboard_ws.title = "Dashboard"
    create_dashboard_sheet(dashboard_ws, content, title, color)
    
    # Detailed data
    data_ws = wb.create_sheet(title="Data")
    populate_sheet(data_ws, content, color, True, True)
    
    # Trends
    if isinstance(content, dict) and kwargs.get('include_charts', True):
        trends_ws = wb.create_sheet(title="Trends")
        create_trends_sheet(trends_ws, content, color)
    
    return wb


def generate_budget_workbook(content, title, color='1B4965', **kwargs):
    """Generate budget planning workbook"""
    
    wb = Workbook()
    
    # Budget sheet
    budget_ws = wb.active
    budget_ws.title = "Budget"
    create_budget_sheet(budget_ws, content, color)
    
    # Actual vs Budget
    comparison_ws = wb.create_sheet(title="Actual vs Budget")
    create_comparison_sheet(comparison_ws, content, color)
    
    # Summary
    summary_ws = wb.create_sheet(title="Summary")
    create_budget_summary(summary_ws, content, color)
    
    return wb


# Sheet creation functions

def populate_sheet(ws, data, color, freeze_panes=True, auto_filter=True):
    """Populate worksheet with data"""
    
    if isinstance(data, dict):
        # Dictionary to table
        if all(isinstance(v, list) for v in data.values()):
            # Column-based data
            headers = list(data.keys())
            
            # Write headers
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                style_header_cell(cell, color)
            
            # Write data
            max_rows = max(len(v) for v in data.values())
            for row_idx in range(max_rows):
                for col_idx, key in enumerate(headers, 1):
                    if row_idx < len(data[key]):
                        value = data[key][row_idx]
                        cell = ws.cell(row=row_idx + 2, column=col_idx, value=value)
                        style_data_cell(cell)
            
            # Adjust column widths
            for col_idx in range(1, len(headers) + 1):
                ws.column_dimensions[get_column_letter(col_idx)].width = 15
            
        else:
            # Key-value pairs
            ws.cell(row=1, column=1, value="Key")
            ws.cell(row=1, column=2, value="Value")
            style_header_cell(ws.cell(row=1, column=1), color)
            style_header_cell(ws.cell(row=1, column=2), color)
            
            for row_idx, (key, value) in enumerate(data.items(), 2):
                cell_key = ws.cell(row=row_idx, column=1, value=key.replace('_', ' ').title())
                cell_value = ws.cell(row=row_idx, column=2, value=value)
                style_data_cell(cell_key)
                style_data_cell(cell_value)
            
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 15
    
    elif isinstance(data, list):
        # List of items
        ws.cell(row=1, column=1, value="Items")
        style_header_cell(ws.cell(row=1, column=1), color)
        
        for row_idx, item in enumerate(data, 2):
            cell = ws.cell(row=row_idx, column=1, value=item)
            style_data_cell(cell)
        
        ws.column_dimensions['A'].width = 25
    
    else:
        # Simple value
        ws.cell(row=1, column=1, value="Data")
        style_header_cell(ws.cell(row=1, column=1), color)
        ws.cell(row=2, column=1, value=str(data))
        style_data_cell(ws.cell(row=2, column=1))
        ws.column_dimensions['A'].width = 25
    
    # Freeze panes
    if freeze_panes:
        ws.freeze_panes = 'A2'
    
    # Auto filter
    if auto_filter and ws.max_row > 1:
        ws.auto_filter.ref = ws.dimensions


def create_summary_sheet(ws, title, data, color):
    """Create summary overview sheet"""
    
    # Title
    ws.merge_cells('A1:D1')
    title_cell = ws['A1']
    title_cell.value = title
    title_cell.font = Font(size=24, bold=True, color=color)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 40
    
    # Date
    ws.merge_cells('A2:D2')
    date_cell = ws['A2']
    date_cell.value = f"Generated: {datetime.now().strftime('%B %d, %Y %H:%M')}"
    date_cell.font = Font(size=12, italic=True)
    date_cell.alignment = Alignment(horizontal='center')
    
    # Key metrics
    row_start = 4
    ws.cell(row=row_start, column=1, value="Key Metrics")
    ws.cell(row=row_start, column=1).font = Font(size=16, bold=True, color=color)
    
    if isinstance(data, dict):
        row = row_start + 2
        for key, value in list(data.items())[:10]:  # Limit to 10 items
            ws.cell(row=row, column=1, value=key.replace('_', ' ').title())
            ws.cell(row=row, column=1).font = Font(bold=True)
            
            value_cell = ws.cell(row=row, column=2, value=value)
            value_cell.font = Font(size=14, color=color)
            
            row += 1
    
    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20


def create_charts_sheet(ws, data, color):
    """Create sheet with various charts"""
    
    # Title
    ws['A1'] = "Data Visualization"
    ws['A1'].font = Font(size=18, bold=True, color=color)
    
    # Add bar chart
    if isinstance(data, dict) and len(data) > 0:
        chart = BarChart()
        chart.title = "Data Overview"
        chart.style = 10
        
        # Create data for chart
        row = 3
        for key, value in list(data.items())[:10]:
            ws.cell(row=row, column=1, value=key.replace('_', ' ').title())
            
            # Handle different value types
            if isinstance(value, (int, float)):
                ws.cell(row=row, column=2, value=value)
            elif isinstance(value, str) and value.replace('.', '').replace('-', '').isdigit():
                ws.cell(row=row, column=2, value=float(value.replace(',', '')))
            else:
                ws.cell(row=row, column=2, value=0)
            
            row += 1
        
        data_range = Reference(ws, min_col=2, min_row=3, max_row=row-1)
        cats = Reference(ws, min_col=1, min_row=3, max_row=row-1)
        
        chart.add_data(data_range, titles_from_data=False)
        chart.set_categories(cats)
        
        ws.add_chart(chart, "D3")


def create_financial_sheet(ws, data, sheet_title, color):
    """Create financial statement sheet"""
    
    # Title
    ws.merge_cells('A1:C1')
    ws['A1'] = sheet_title
    ws['A1'].font = Font(size=20, bold=True, color=color)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 35
    
    # Headers
    ws['A3'] = "Account"
    ws['B3'] = "Amount"
    ws['C3'] = "Percentage"
    
    for cell in [ws['A3'], ws['B3'], ws['C3']]:
        style_header_cell(cell, color)
    
    # Data
    if isinstance(data, dict):
        row = 4
        total = sum(v for v in data.values() if isinstance(v, (int, float)))
        
        for key, value in data.items():
            ws.cell(row=row, column=1, value=key.replace('_', ' ').title())
            
            if isinstance(value, (int, float)):
                amount_cell = ws.cell(row=row, column=2, value=value)
                amount_cell.number_format = '$#,##0.00'
                
                if total > 0:
                    pct_cell = ws.cell(row=row, column=3, value=value/total)
                    pct_cell.number_format = '0.00%'
            else:
                ws.cell(row=row, column=2, value=value)
            
            row += 1
        
        # Total row
        total_row = row + 1
        ws.cell(row=total_row, column=1, value="TOTAL")
        ws.cell(row=total_row, column=1).font = Font(bold=True)
        total_cell = ws.cell(row=total_row, column=2, value=total)
        total_cell.number_format = '$#,##0.00'
        total_cell.font = Font(bold=True)
        
        # Add border to total row
        for col in range(1, 4):
            cell = ws.cell(row=total_row, column=col)
            cell.border = Border(top=Side(style='double'))
    
    # Column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12


def create_financial_summary(ws, data, color):
    """Create financial summary with key ratios"""
    
    create_summary_sheet(ws, "Financial Summary", data, color)


def create_financial_charts(ws, data, color):
    """Create financial analysis charts"""
    
    create_charts_sheet(ws, data, color)


def create_inventory_sheet(ws, data, color):
    """Create inventory tracking sheet"""
    
    # Headers
    headers = ["Item", "Quantity", "Unit Price", "Total Value", "Reorder Level", "Status"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        style_header_cell(cell, color)
    
    # Sample data structure
    if isinstance(data, dict):
        row = 2
        for key, value in data.items():
            ws.cell(row=row, column=1, value=key.replace('_', ' ').title())
            
            if isinstance(value, (int, float)):
                ws.cell(row=row, column=2, value=value)
                ws.cell(row=row, column=3, value=10.00)
                
                total_cell = ws.cell(row=row, column=4, value=value * 10)
                total_cell.number_format = '$#,##0.00'
                
                ws.cell(row=row, column=5, value=50)
                
                status = "Low" if value < 50 else "OK"
                status_cell = ws.cell(row=row, column=6, value=status)
                
                if status == "Low":
                    status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                else:
                    status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            
            row += 1
    
    # Column widths
    for col_idx in range(1, 7):
        ws.column_dimensions[get_column_letter(col_idx)].width = 15
    
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions


def create_alert_sheet(ws, data, color):
    """Create low stock alert sheet"""
    
    ws['A1'] = "Low Stock Items"
    ws['A1'].font = Font(size=18, bold=True, color="FF0000")
    
    # Headers
    ws.cell(row=3, column=1, value="Item")
    ws.cell(row=3, column=2, value="Current Stock")
    ws.cell(row=3, column=3, value="Reorder Level")
    
    for col in range(1, 4):
        style_header_cell(ws.cell(row=3, column=col), color)
    
    # Column widths
    for col_idx in range(1, 4):
        ws.column_dimensions[get_column_letter(col_idx)].width = 20


def create_inventory_stats(ws, data, color):
    """Create inventory statistics sheet"""
    
    create_summary_sheet(ws, "Inventory Statistics", data, color)


def create_dashboard_sheet(ws, title, data, color):
    """Create KPI dashboard sheet"""
    
    # Title
    ws.merge_cells('A1:F1')
    ws['A1'] = f"📊 {title}"
    ws['A1'].font = Font(size=28, bold=True, color=color)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 50
    
    # Date
    ws.merge_cells('A2:F2')
    ws['A2'] = datetime.now().strftime('%B %d, %Y - %H:%M')
    ws['A2'].alignment = Alignment(horizontal='center')
    ws['A2'].font = Font(size=12, italic=True)
    
    # KPI Cards
    if isinstance(data, dict):
        row = 4
        col = 1
        
        for i, (key, value) in enumerate(list(data.items())[:6]):
            # Create KPI card
            ws.merge_cells(start_row=row, start_column=col, end_row=row+2, end_column=col+1)
            
            # Title
            title_cell = ws.cell(row=row, column=col)
            title_cell.value = key.replace('_', ' ').title()
            title_cell.font = Font(size=12, bold=True)
            title_cell.alignment = Alignment(horizontal='center', vertical='top')
            
            # Value
            value_cell = ws.cell(row=row+1, column=col)
            value_cell.value = value
            value_cell.font = Font(size=24, bold=True, color=color)
            value_cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Background
            fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
            for r in range(row, row+3):
                for c in range(col, col+2):
                    ws.cell(row=r, column=c).fill = fill
                    ws.cell(row=r, column=c).border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
            
            # Move to next position
            col += 3
            if col > 6:
                col = 1
                row += 5


def create_trends_sheet(ws, data, color):
    """Create trends analysis sheet"""
    
    create_charts_sheet(ws, data, color)


def create_budget_sheet(ws, data, color):
    """Create budget planning sheet"""
    
    # Headers
    headers = ["Category", "Budgeted", "Actual", "Variance", "Variance %"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        style_header_cell(cell, color)
    
    if isinstance(data, dict):
        row = 2
        for key, value in data.items():
            ws.cell(row=row, column=1, value=key.replace('_', ' ').title())
            
            if isinstance(value, (int, float)):
                # Budgeted
                budget_cell = ws.cell(row=row, column=2, value=value)
                budget_cell.number_format = '$#,##0.00'
                
                # Actual (sample data)
                actual = value * 0.95  # 95% of budget
                actual_cell = ws.cell(row=row, column=3, value=actual)
                actual_cell.number_format = '$#,##0.00'
                
                # Variance
                variance = actual - value
                variance_cell = ws.cell(row=row, column=4, value=variance)
                variance_cell.number_format = '$#,##0.00'
                
                # Variance %
                if value != 0:
                    var_pct = variance / value
                    var_pct_cell = ws.cell(row=row, column=5, value=var_pct)
                    var_pct_cell.number_format = '0.00%'
                    
                    # Color code
                    if var_pct < -0.1:
                        var_pct_cell.font = Font(color="FF0000")
                    elif var_pct > 0:
                        var_pct_cell.font = Font(color="00B050")
            
            row += 1
    
    # Column widths
    for col_idx in range(1, 6):
        ws.column_dimensions[get_column_letter(col_idx)].width = 15
    
    ws.freeze_panes = 'A2'


def create_comparison_sheet(ws, data, color):
    """Create actual vs budget comparison"""
    
    create_budget_sheet(ws, data, color)


def create_budget_summary(ws, data, color):
    """Create budget summary sheet"""
    
    create_summary_sheet(ws, "Budget Summary", data, color)


def add_basic_chart(ws, data):
    """Add a basic chart to worksheet"""
    
    if not isinstance(data, dict) or len(data) == 0:
        return
    
    # Pie chart
    chart = PieChart()
    chart.title = "Data Distribution"
    chart.style = 10
    
    # Use existing data
    max_row = ws.max_row
    
    labels = Reference(ws, min_col=1, min_row=2, max_row=max_row)
    data_ref = Reference(ws, min_col=2, min_row=1, max_row=max_row)
    
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(labels)
    
    # Add chart to sheet
    ws.add_chart(chart, f"E2")


# Styling functions

def style_header_cell(cell, color):
    """Apply header cell styling"""
    cell.font = Font(bold=True, color='FFFFFF', size=12)
    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )


def style_data_cell(cell):
    """Apply data cell styling"""
    cell.alignment = Alignment(horizontal='left', vertical='center')
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
